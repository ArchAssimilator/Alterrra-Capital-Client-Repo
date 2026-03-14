from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = Workbook()

# Styles
input_font = Font(color='0000FF')
calc_font = Font(color='000000')
green_font = Font(color='008000')
bold_input = Font(bold=True, color='0000FF')
bold_calc = Font(bold=True, color='000000')
bold_green = Font(bold=True, color='008000')
section_font = Font(bold=True, size=11)
col_header_font = Font(bold=True)
col_header_fill = PatternFill('solid', fgColor='D9E1F2')
pct_fmt = '0.0%'
num_fmt = '#,##0.0'
int_fmt = '#,##0'
mult_fmt = '0.0"x"'
dollar_fmt = '$#,##0.0'
dollar_int = '$#,##0'

def si(ws, r, c, v, fmt=None, bold=False):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = bold_input if bold else input_font
    if fmt: cell.number_format = fmt
    return cell

def sc(ws, r, c, f, fmt=None, bold=False):
    cell = ws.cell(row=r, column=c, value=f)
    cell.font = bold_calc if bold else calc_font
    if fmt: cell.number_format = fmt
    return cell

def sx(ws, r, c, f, fmt=None, bold=False):
    cell = ws.cell(row=r, column=c, value=f)
    cell.font = bold_green if bold else green_font
    if fmt: cell.number_format = fmt
    return cell

def sl(ws, r, c, v, bold=False, section=False):
    cell = ws.cell(row=r, column=c, value=v)
    if section: cell.font = section_font
    elif bold: cell.font = Font(bold=True)
    return cell

def set_headers(ws, row, labels, start_col=1):
    for i, label in enumerate(labels):
        cell = ws.cell(row=row, column=start_col+i, value=label)
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.alignment = Alignment(horizontal='center')

def col(c):
    return get_column_letter(c)

# ============================================================
# SHEET 1: Exhibit 1 - Portfolio Companies (ALL INPUT)
# ============================================================
ws1 = wb.active
ws1.title = "Ex1 Portfolio"
ws1.column_dimensions['A'].width = 42
ws1.column_dimensions['B'].width = 16
ws1.column_dimensions['C'].width = 22

sl(ws1, 1, 1, "Exhibit 1 — Roark Capital Restaurant and Food Portfolio Companies", section=True)
sl(ws1, 2, 1, "As of October 2017")
set_headers(ws1, 3, ["Name", "Year Acquired", "Exit Year"])

portfolio = [
    ("Carvel", 2001, "—"), ("Cinnabon", 2004, "—"), ("Seattle's Best Coffee", 2004, "—"),
    ("McAlister's Deli", 2005, "—"), ("Schlotzsky's Bakery & Cafe", 2006, "—"),
    ("Moe's Southwest Grill", 2007, "—"), ("Auntie Anne's", 2010, "—"),
    ("Wingstop", 2010, "2015 IPO"), ("Arby's", 2011, "—"),
    ("Corner Bakery Cafe", 2011, "—"), ("Il Fornaio Italian Restaurant & Bakery", 2011, "—"),
    ("Carl's Jr. Hardee's", 2013, "—"), ("Miller's Ale House", 2013, "—"),
    ("Naf Naf Grill", 2015, "—"), ("Jimmy John's", 2016, "—"),
    ("Culver's", 2017, "—"), ("Jim 'N Nick's Community Bar-B-Q", 2017, "—"),
]
for i, (name, year, exit_yr) in enumerate(portfolio):
    r = 4 + i
    sl(ws1, r, 1, name)
    si(ws1, r, 2, year, fmt='0')
    si(ws1, r, 3, exit_yr)

# ============================================================
# SHEET 2: Exhibit 2 - Historical Income Statement
# ============================================================
ws2 = wb.create_sheet("Ex2 Income Statement")
ws2.column_dimensions['A'].width = 42
for c in range(2, 8):
    ws2.column_dimensions[col(c)].width = 14

sl(ws2, 1, 1, "Exhibit 2 — BWW's Historical Income Statement", section=True)
sl(ws2, 2, 1, "$ in millions | Source: F-1984.PDF, Page 11")
set_headers(ws2, 3, ["Line Item", "2012", "2013", "2014", "2015", "2016", "2017E"])

# Row 4: Restaurant sales (INPUT)
sl(ws2, 4, 1, "Restaurant sales (company revenues)")
for i, v in enumerate([964.0, 1185.3, 1423.0, 1715.0, 1891.6, 1977.5]):
    si(ws2, 4, 2+i, v, fmt=num_fmt)

# Row 5: YoY growth (2012=INPUT no prior, 2013+=CALC)
sl(ws2, 5, 1, "  YoY growth %")
si(ws2, 5, 2, 0.344, fmt=pct_fmt)
for c in range(3, 8):
    sc(ws2, 5, c, f"={col(c)}4/{col(c-1)}4-1", fmt=pct_fmt)

# Row 6: Franchise royalties (INPUT)
sl(ws2, 6, 1, "Franchise royalties and fees")
for i, v in enumerate([76.5, 81.4, 93.2, 97.7, 95.2, 100.3]):
    si(ws2, 6, 2+i, v, fmt=num_fmt)

# Row 7: YoY growth franchise
sl(ws2, 7, 1, "  YoY growth %")
si(ws2, 7, 2, 0.141, fmt=pct_fmt)
for c in range(3, 8):
    sc(ws2, 7, c, f"={col(c)}6/{col(c-1)}6-1", fmt=pct_fmt)

# Row 8: Total revenue (CALCULATED)
sl(ws2, 8, 1, "Total revenue", bold=True)
for c in range(2, 8):
    sc(ws2, 8, c, f"={col(c)}4+{col(c)}6", fmt=dollar_fmt, bold=True)

# Row 9: YoY growth total
sl(ws2, 9, 1, "  YoY growth %")
si(ws2, 9, 2, 0.326, fmt=pct_fmt)
for c in range(3, 8):
    sc(ws2, 9, c, f"={col(c)}8/{col(c-1)}8-1", fmt=pct_fmt)

# Row 11: Cost of sales (INPUT)
sl(ws2, 11, 1, "Cost of sales")
for i, v in enumerate([303.7, 363.8, 413.9, 507.8, 564.7, 617.0]):
    si(ws2, 11, 2+i, v, fmt=num_fmt)

# Row 12: Cost of sales %
sl(ws2, 12, 1, "  Cost of sales/company restaurant sales")
for c in range(2, 8):
    sc(ws2, 12, c, f"={col(c)}11/{col(c)}4", fmt=pct_fmt)

# Row 13: Labor costs (INPUT)
sl(ws2, 13, 1, "Labor costs")
for i, v in enumerate([289.2, 360.3, 444.2, 542.8, 599.0, 629.3]):
    si(ws2, 13, 2+i, v, fmt=num_fmt)

# Row 14: Labor %
sl(ws2, 14, 1, "  Labor/company restaurant sales")
for c in range(2, 8):
    sc(ws2, 14, c, f"={col(c)}13/{col(c)}4", fmt=pct_fmt)

# Row 15: Operating costs (INPUT)
sl(ws2, 15, 1, "Operating costs")
for i, v in enumerate([141.4, 174.3, 209.6, 250.8, 285.1, 303.1]):
    si(ws2, 15, 2+i, v, fmt=num_fmt)

# Row 16: Operating costs %
sl(ws2, 16, 1, "  Operating costs/company restaurant sales")
for c in range(2, 8):
    sc(ws2, 16, c, f"={col(c)}15/{col(c)}4", fmt=pct_fmt)

# Row 17: Occupancy costs (INPUT)
sl(ws2, 17, 1, "Occupancy costs")
for i, v in enumerate([54.1, 68.4, 78.9, 94.6, 108.9, 114.2]):
    si(ws2, 17, 2+i, v, fmt=num_fmt)

# Row 18: Occupancy %
sl(ws2, 18, 1, "  Occupancy costs/company restaurant sales")
for c in range(2, 8):
    sc(ws2, 18, c, f"={col(c)}17/{col(c)}4", fmt=pct_fmt)

# Row 19: Gross profit (CALCULATED)
sl(ws2, 19, 1, "Gross profit", bold=True)
for c in range(2, 8):
    sc(ws2, 19, c, f"={col(c)}8-{col(c)}11-{col(c)}13-{col(c)}15-{col(c)}17", fmt=dollar_fmt, bold=True)

# Row 20: Gross profit margin
sl(ws2, 20, 1, "  Gross profit margin")
for c in range(2, 8):
    sc(ws2, 20, c, f"={col(c)}19/{col(c)}8", fmt=pct_fmt)

# Row 22: SG&A (INPUT)
sl(ws2, 22, 1, "Selling, general, and administrative (SG&A)")
for i, v in enumerate([84.1, 96.2, 118.0, 129.1, 123.1, 133.1]):
    si(ws2, 22, 2+i, v, fmt=num_fmt)

# Row 23: SG&A %
sl(ws2, 23, 1, "  SG&A/total revenue")
for c in range(2, 8):
    sc(ws2, 23, c, f"={col(c)}22/{col(c)}8", fmt=pct_fmt)

# Row 24: Pre-opening costs (INPUT)
sl(ws2, 24, 1, "Pre-opening costs")
for i, v in enumerate([14.6, 14.6, 13.5, 14.2, 8.7, 4.3]):
    si(ws2, 24, 2+i, v, fmt=num_fmt)

# Row 25: D&A (INPUT)
sl(ws2, 25, 1, "Depreciation and amortization")
for i, v in enumerate([67.5, 85.0, 98.5, 127.5, 152.1, 151.4]):
    si(ws2, 25, 2+i, v, fmt=num_fmt)

# Row 26: Other operating expenses total (CALCULATED)
sl(ws2, 26, 1, "Other operating expenses, total", bold=True)
for c in range(2, 8):
    sc(ws2, 26, c, f"={col(c)}22+{col(c)}24+{col(c)}25", fmt=dollar_fmt, bold=True)

# Row 28: Operating income (CALCULATED)
sl(ws2, 28, 1, "Operating income", bold=True)
for c in range(2, 8):
    sc(ws2, 28, c, f"={col(c)}19-{col(c)}26", fmt=dollar_fmt, bold=True)

# Row 29: Operating income %
sl(ws2, 29, 1, "  Operating income/total revenue")
for c in range(2, 8):
    sc(ws2, 29, c, f"={col(c)}28/{col(c)}8", fmt=pct_fmt)

# ============================================================
# SHEET 3: Exhibit 3 - Historical Balance Sheet
# ============================================================
ws3 = wb.create_sheet("Ex3 Balance Sheet")
ws3.column_dimensions['A'].width = 42
for c in range(2, 5):
    ws3.column_dimensions[col(c)].width = 18

sl(ws3, 1, 1, "Exhibit 3 — BWW's Historical Balance Sheet", section=True)
sl(ws3, 2, 1, "$ in millions | Source: F-1984.PDF, Page 12")
set_headers(ws3, 3, ["Line Item", "Dec. 30, 2015", "Dec. 30, 2016", "Sept. 30, 2017"])

sl(ws3, 4, 1, "ASSETS", bold=True)
ca_items = [
    ("Cash and equivalents", [20.3, 49.3, 30.7]),
    ("Accounts receivable", [34.1, 34.2, 42.4]),
    ("Other receivables", [21.6, 1.0, 2.1]),
    ("Inventory", [15.4, 16.5, 14.7]),
    ("Prepaid expenses", [6.4, 9.1, 10.1]),
    ("Other current assets", [100.1, 66.5, 23.3]),
]
for i, (label, vals) in enumerate(ca_items):
    r = 5 + i
    sl(ws3, r, 1, label)
    for j, v in enumerate(vals):
        si(ws3, r, 2+j, v, fmt=num_fmt)

sl(ws3, 11, 1, "Total current assets", bold=True)
for c in range(2, 5):
    sc(ws3, 11, c, f"=SUM({col(c)}5:{col(c)}10)", fmt=num_fmt, bold=True)

sl(ws3, 12, 1, "Gross property, plant, and equipment")
for i, v in enumerate([1059.5, 1150.5, 1179.6]):
    si(ws3, 12, 2+i, v, fmt=num_fmt)

sl(ws3, 13, 1, "Accumulated depreciation")
for i, v in enumerate([-454.7, -557.7, -637.9]):
    si(ws3, 13, 2+i, v, fmt='#,##0.0;(#,##0.0)')

sl(ws3, 14, 1, "Net property, plant, and equipment", bold=True)
for c in range(2, 5):
    sc(ws3, 14, c, f"={col(c)}12+{col(c)}13", fmt=num_fmt, bold=True)

for i, (label, vals) in enumerate([
    ("Goodwill", [114.1, 117.2, 117.2]),
    ("Other intangibles", [138.2, 129.4, 109.0]),
    ("Other long-term assets", [17.7, 31.2, 39.2]),
]):
    r = 15 + i
    sl(ws3, r, 1, label)
    for j, v in enumerate(vals):
        si(ws3, r, 2+j, v, fmt=num_fmt)

sl(ws3, 18, 1, "Total assets", bold=True)
for c in range(2, 5):
    sc(ws3, 18, c, f"={col(c)}11+{col(c)}14+{col(c)}15+{col(c)}16+{col(c)}17", fmt=num_fmt, bold=True)

sl(ws3, 20, 1, "LIABILITIES", bold=True)
cl_items = [
    ("Accounts payable", [44.8, 45.8, 40.9]),
    ("Accrued expenses", [99.2, 104.0, 65.0]),
    ("Current portion of long-term debt", [2.1, 3.7, 4.6]),
    ("Unearned revenue, current", [81.3, 87.6, 6.9]),
    ("Other current liabilities", [36.2, 0.9, 59.3]),
]
for i, (label, vals) in enumerate(cl_items):
    r = 21 + i
    sl(ws3, r, 1, label)
    for j, v in enumerate(vals):
        si(ws3, r, 2+j, v, fmt=num_fmt)

sl(ws3, 26, 1, "Total current liabilities", bold=True)
for c in range(2, 5):
    sc(ws3, 26, c, f"=SUM({col(c)}21:{col(c)}25)", fmt=num_fmt, bold=True)

for i, (label, vals) in enumerate([
    ("Long-term debt", [71.0, 205.3, 420.4]),
    ("Deferred tax liability, non-current", [23.7, 21.6, 12.9]),
    ("Other non-current liabilities", [58.3, 60.5, 57.6]),
]):
    r = 27 + i
    sl(ws3, r, 1, label)
    for j, v in enumerate(vals):
        si(ws3, r, 2+j, v, fmt=num_fmt)

sl(ws3, 30, 1, "Total liabilities", bold=True)
for c in range(2, 5):
    sc(ws3, 30, c, f"={col(c)}26+{col(c)}27+{col(c)}28+{col(c)}29", fmt=num_fmt, bold=True)

for i, (label, vals, fmt) in enumerate([
    ("Common stock", [160.4, 147.2, 142.7], num_fmt),
    ("Retained earnings", [499.1, 374.7, 124.3], num_fmt),
    ("Comprehensive income and other", [-4.1, -3.9, -3.6], '#,##0.0;(#,##0.0)'),
]):
    r = 31 + i
    sl(ws3, r, 1, label)
    for j, v in enumerate(vals):
        si(ws3, r, 2+j, v, fmt=fmt)

sl(ws3, 34, 1, "Total common equity", bold=True)
for c in range(2, 5):
    sc(ws3, 34, c, f"={col(c)}31+{col(c)}32+{col(c)}33", fmt=num_fmt, bold=True)

sl(ws3, 35, 1, "Minority interest")
for i, v in enumerate([0.4, -0.1, -0.6]):
    si(ws3, 35, 2+i, v, fmt='#,##0.0;(#,##0.0)')

sl(ws3, 36, 1, "Total equity", bold=True)
for c in range(2, 5):
    sc(ws3, 36, c, f"={col(c)}34+{col(c)}35", fmt=num_fmt, bold=True)

sl(ws3, 37, 1, "Total liabilities and equity", bold=True)
for c in range(2, 5):
    sc(ws3, 37, c, f"={col(c)}30+{col(c)}36", fmt=num_fmt, bold=True)

# ============================================================
# SHEET 4: Exhibit 8 - Stock Return Performance (ALL INPUT)
# ============================================================
ws8 = wb.create_sheet("Ex8 Stock Returns")
ws8.column_dimensions['A'].width = 35
for c in range(2, 6):
    ws8.column_dimensions[col(c)].width = 14

sl(ws8, 1, 1, "Exhibit 8 — Stock Return Performance", section=True)
sl(ws8, 2, 1, "Cumulative return per $1 invested through Dec 2016 | Source: Page 17")
set_headers(ws8, 3, ["Index", "Five-year", "Three-year", "Two-year", "One-year"])

for i, (label, vals) in enumerate([
    ("BWW's stock", [2.32, 1.05, 0.87, 0.97]),
    ("S&P 500 Index", [1.71, 1.21, 1.17, 1.10]),
    ("S&P 1500 Restaurant Index", [1.31, 1.28, 1.20, 1.04]),
]):
    r = 4 + i
    sl(ws8, r, 1, label)
    for j, v in enumerate(vals):
        si(ws8, r, 2+j, v, fmt='0.00')

# ============================================================
# SHEET 5: Exhibit 9 - Performance Metrics
# ============================================================
ws9 = wb.create_sheet("Ex9 Performance")
ws9.column_dimensions['A'].width = 50
for c in range(2, 8):
    ws9.column_dimensions[col(c)].width = 14

sl(ws9, 1, 1, "Exhibit 9 — BWW's Performance Metrics", section=True)
sl(ws9, 2, 1, "Source: F-1984.PDF, Page 18")
set_headers(ws9, 3, ["Metric", "2012", "2013", "2014", "2015", "2016", "Q3 2017A"])

sl(ws9, 4, 1, "Revenue growth", section=True)
for i, (label, vals) in enumerate([
    ("Company operated", [0.344, 0.230, 0.200, 0.205, 0.103, 0.040]),
    ("Franchised", [0.141, 0.063, 0.146, 0.048, -0.026, 0.046]),
    ("Total restaurant", [0.326, 0.217, 0.197, 0.196, 0.096, 0.041]),
]):
    r = 5 + i
    sl(ws9, r, 1, label)
    for j, v in enumerate(vals):
        si(ws9, r, 2+j, v, fmt=pct_fmt)

sl(ws9, 9, 1, "Same-store sales growth", section=True)
for i, (label, vals) in enumerate([
    ("Company operated", [0.066, 0.039, 0.065, 0.043, -0.024, -0.023]),
    ("Franchised", [0.065, 0.033, 0.056, 0.025, -0.026, -0.032]),
]):
    r = 10 + i
    sl(ws9, r, 1, label)
    for j, v in enumerate(vals):
        si(ws9, r, 2+j, v, fmt=pct_fmt)

sl(ws9, 13, 1, "Margins", section=True)
for i, (label, vals) in enumerate([
    ('Restaurant level ("four-wall") margin', [0.182, 0.184, 0.194, 0.186, 0.177, 0.162]),
    ("Operating margin", [0.083, 0.082, 0.092, 0.081, 0.073, 0.057]),
]):
    r = 14 + i
    sl(ws9, r, 1, label)
    for j, v in enumerate(vals):
        si(ws9, r, 2+j, v, fmt=pct_fmt)

sl(ws9, 17, 1, "Store count", section=True)
sl(ws9, 18, 1, "Company operated")
for i, v in enumerate([381, 434, 491, 596, 631, 638]):
    si(ws9, 18, 2+i, v, fmt=int_fmt)

sl(ws9, 19, 1, "  YoY growth %")
si(ws9, 19, 2, 0.194, fmt=pct_fmt)
for c in range(3, 8):
    sc(ws9, 19, c, f"={col(c)}18/{col(c-1)}18-1", fmt=pct_fmt)

sl(ws9, 20, 1, "Franchised")
for i, v in enumerate([510, 559, 591, 579, 609, 633]):
    si(ws9, 20, 2+i, v, fmt=int_fmt)

sl(ws9, 21, 1, "  YoY growth %")
si(ws9, 21, 2, 0.024, fmt=pct_fmt)
for c in range(3, 8):
    sc(ws9, 21, c, f"={col(c)}20/{col(c-1)}20-1", fmt=pct_fmt)

sl(ws9, 22, 1, "Total restaurants")
for c in range(2, 8):
    sc(ws9, 22, c, f"={col(c)}18+{col(c)}20", fmt=int_fmt)

sl(ws9, 23, 1, "  YoY growth %")
si(ws9, 23, 2, 0.091, fmt=pct_fmt)
for c in range(3, 8):
    sc(ws9, 23, c, f"={col(c)}22/{col(c-1)}22-1", fmt=pct_fmt)

sl(ws9, 25, 1, "Franchising mix", bold=True)
for c in range(2, 8):
    sc(ws9, 25, c, f"={col(c)}20/{col(c)}22", fmt=pct_fmt)

sl(ws9, 27, 1, "Average unit volume (AUV)", section=True)
sl(ws9, 28, 1, "Company operated AUV ($K, annual)")
for i, v in enumerate([2793, 2932, 3148, 3258, 3140, 3038]):
    si(ws9, 28, 2+i, v, fmt=dollar_int)

sl(ws9, 29, 1, "Franchised AUV ($K, annual)")
for i, v in enumerate([2940, 3099, 3266, 3352, 3263, 3135]):
    si(ws9, 29, 2+i, v, fmt=dollar_int)

sl(ws9, 31, 1, "Avg weekly sales per company-operated unit")
for i, v in enumerate([53783, 56377, 60470, 62529, 60366, 57930]):
    si(ws9, 31, 2+i, v, fmt=dollar_int)

sl(ws9, 32, 1, "Avg weekly sales per franchised unit")
for i, v in enumerate([56489, 59582, 62596, 64474, 62662, 59964]):
    si(ws9, 32, 2+i, v, fmt=dollar_int)

# ============================================================
# SHEET 6: Exhibit 10 - LBO Assumptions
# ============================================================
ws10a = wb.create_sheet("Ex10 Assumptions")
ws10a.column_dimensions['A'].width = 50
for c in range(2, 8):
    ws10a.column_dimensions[col(c)].width = 16

sl(ws10a, 1, 1, "Exhibit 10 — LBO Forecast Model Assumptions", section=True)
sl(ws10a, 2, 1, "$ in millions (unless otherwise noted) | Source: Page 19")
set_headers(ws10a, 3, ["Line Item", "Closing 2017E", "2018P", "2019P", "2020P", "2021P", "2022P"])

sl(ws10a, 5, 1, "Store Count — Company", section=True)

sl(ws10a, 6, 1, "Beginning balance, company restaurants")
for c in range(3, 8):
    sc(ws10a, 6, c, f"={col(c-1)}9", fmt=int_fmt)

sl(ws10a, 7, 1, "  Less: Refranchised restaurants")
for i, v in enumerate([None, 33, 40, 40, 40, 20]):
    if v is not None:
        si(ws10a, 7, 2+i, v, fmt=int_fmt)

sl(ws10a, 8, 1, "  Plus: New company restaurants")
for i in range(1, 6):
    si(ws10a, 8, 2+i, 0, fmt=int_fmt)

sl(ws10a, 9, 1, "Ending balance, company restaurants", bold=True)
si(ws10a, 9, 2, 640, fmt=int_fmt, bold=True)
for c in range(3, 8):
    sc(ws10a, 9, c, f"={col(c)}6-{col(c)}7+{col(c)}8", fmt=int_fmt, bold=True)

sl(ws10a, 10, 1, "  Unit growth")
for c in range(3, 8):
    sc(ws10a, 10, c, f"={col(c)}9/{col(c-1)}9-1", fmt=pct_fmt)

sl(ws10a, 12, 1, "Store Count — Franchised", section=True)

sl(ws10a, 13, 1, "Beginning balance, franchised restaurants")
for c in range(3, 8):
    sc(ws10a, 13, c, f"={col(c-1)}16", fmt=int_fmt)

sl(ws10a, 14, 1, "  Plus: Refranchised restaurants")
for c in range(3, 8):
    sx(ws10a, 14, c, f"={col(c)}7", fmt=int_fmt)

sl(ws10a, 15, 1, "  Plus: New franchised restaurants")
for i, v in enumerate([None, 40, 35, 30, 25, 25]):
    if v is not None:
        si(ws10a, 15, 2+i, v, fmt=int_fmt)

sl(ws10a, 16, 1, "Ending balance, franchised restaurants", bold=True)
si(ws10a, 16, 2, 647, fmt=int_fmt, bold=True)
for c in range(3, 8):
    sc(ws10a, 16, c, f"={col(c)}13+{col(c)}14+{col(c)}15", fmt=int_fmt, bold=True)

sl(ws10a, 17, 1, "  Unit growth")
for c in range(3, 8):
    sc(ws10a, 17, c, f"={col(c)}16/{col(c-1)}16-1", fmt=pct_fmt)

sl(ws10a, 19, 1, "Averages", section=True)
sl(ws10a, 20, 1, "Average company restaurants")
for c in range(3, 8):
    sc(ws10a, 20, c, f"=({col(c)}6+{col(c)}9)/2", fmt=num_fmt)

sl(ws10a, 21, 1, "Average franchised restaurants")
for c in range(3, 8):
    sc(ws10a, 21, c, f"=({col(c)}13+{col(c)}16)/2", fmt=num_fmt)

sl(ws10a, 23, 1, "Weekly Sales", section=True)
sl(ws10a, 24, 1, "Avg weekly sales per company restaurant ($M)")
for i, v in enumerate([None, 0.0604, 0.0612, 0.0621, 0.0630, 0.0643]):
    if v is not None:
        si(ws10a, 24, 2+i, v, fmt='0.0000')

sl(ws10a, 25, 1, "Avg weekly sales per franchised restaurant ($M)")
for i, v in enumerate([None, 0.0627, 0.0636, 0.0645, 0.0655, 0.0668]):
    if v is not None:
        si(ws10a, 25, 2+i, v, fmt='0.0000')

sl(ws10a, 27, 1, "Key Assumptions", section=True)
sl(ws10a, 28, 1, "Franchise royalties")
for i in range(1, 6):
    si(ws10a, 28, 2+i, 0.05, fmt=pct_fmt)

sl(ws10a, 29, 1, "Cost of sales/company restaurant sales")
for i, v in enumerate([0.312, 0.312, 0.312, 0.309, 0.306, 0.304]):
    si(ws10a, 29, 2+i, v, fmt=pct_fmt)

sl(ws10a, 30, 1, "Labor/company restaurant sales")
for i, v in enumerate([None, 0.318, 0.316, 0.314, 0.312, 0.310]):
    if v is not None:
        si(ws10a, 30, 2+i, v, fmt=pct_fmt)

sl(ws10a, 31, 1, "Operating costs/company restaurant sales")
for i, v in enumerate([0.153, 0.150, 0.149, 0.148, 0.147, 0.146]):
    si(ws10a, 31, 2+i, v, fmt=pct_fmt)

sl(ws10a, 32, 1, "Occupancy costs/company restaurant sales")
for i, v in enumerate([None, 0.058, 0.058, 0.058, 0.056, 0.054]):
    if v is not None:
        si(ws10a, 32, 2+i, v, fmt=pct_fmt)

sl(ws10a, 33, 1, "SG&A/total revenues")
for i, v in enumerate([0.064, 0.065, 0.064, 0.063, 0.063, 0.062]):
    si(ws10a, 33, 2+i, v, fmt=pct_fmt)

sl(ws10a, 35, 1, "D&A per company restaurant ($M)")
for i in range(1, 6):
    si(ws10a, 35, 2+i, 0.24, fmt='0.00')

sl(ws10a, 37, 1, "Expected tax rate")
for i in range(1, 6):
    si(ws10a, 37, 2+i, 0.25, fmt=pct_fmt)

sl(ws10a, 39, 1, "After-tax sale price of refranchised restaurants ($M)")
for i in range(1, 6):
    si(ws10a, 39, 2+i, 1.8, fmt='0.0')

# ============================================================
# SHEET 7: Exhibit 10 - Projected Income Statement
# ============================================================
ws10b = wb.create_sheet("Ex10 Projected IS")
ws10b.column_dimensions['A'].width = 50
for c in range(2, 8):
    ws10b.column_dimensions[col(c)].width = 16

sl(ws10b, 1, 1, "Exhibit 10 (continued) — BWW's Projected Income Statement", section=True)
sl(ws10b, 2, 1, "$ in millions | Source: Page 20")
set_headers(ws10b, 3, ["Line Item", "Closing 2017E", "2018P", "2019P", "2020P", "2021P", "2022P"])

AS = "'Ex10 Assumptions'"
IS = "'Ex2 Income Statement'"

sl(ws10b, 5, 1, "Revenue", section=True)

sl(ws10b, 6, 1, "Average number of company restaurants")
for c in range(3, 8):
    sx(ws10b, 6, c, f"={AS}!{col(c)}20", fmt=num_fmt)

sl(ws10b, 7, 1, "Avg weekly sales per company restaurant ($M)")
for c in range(3, 8):
    sx(ws10b, 7, c, f"={AS}!{col(c)}24", fmt='0.0000')

sl(ws10b, 8, 1, "Company revenue")
sx(ws10b, 8, 2, f"={IS}!G4", fmt=num_fmt)
for c in range(3, 8):
    sc(ws10b, 8, c, f"={col(c)}6*{col(c)}7*52", fmt=num_fmt)

sl(ws10b, 10, 1, "Average number of franchised restaurants")
for c in range(3, 8):
    sx(ws10b, 10, c, f"={AS}!{col(c)}21", fmt=num_fmt)

sl(ws10b, 11, 1, "Avg weekly sales per franchised restaurant ($M)")
for c in range(3, 8):
    sx(ws10b, 11, c, f"={AS}!{col(c)}25", fmt='0.0000')

sl(ws10b, 12, 1, "Franchise sales")
for c in range(3, 8):
    sc(ws10b, 12, c, f"={col(c)}10*{col(c)}11*52", fmt=num_fmt)

sl(ws10b, 13, 1, "Royalty revenue")
sx(ws10b, 13, 2, f"={IS}!G6", fmt=num_fmt)
for c in range(3, 8):
    sc(ws10b, 13, c, f"={col(c)}12*{AS}!{col(c)}28", fmt=num_fmt)

sl(ws10b, 14, 1, "Total revenue", bold=True)
sc(ws10b, 14, 2, "=B8+B13", fmt=dollar_fmt, bold=True)
for c in range(3, 8):
    sc(ws10b, 14, c, f"={col(c)}8+{col(c)}13", fmt=dollar_fmt, bold=True)

sl(ws10b, 16, 1, "Expenses", section=True)

sl(ws10b, 17, 1, "Cost of sales")
for c in range(3, 8):
    sc(ws10b, 17, c, f"={col(c)}8*{AS}!{col(c)}29", fmt=num_fmt)

sl(ws10b, 18, 1, "Labor")
for c in range(3, 8):
    sc(ws10b, 18, c, f"={col(c)}8*{AS}!{col(c)}30", fmt=num_fmt)

sl(ws10b, 19, 1, "Operating costs")
for c in range(3, 8):
    sc(ws10b, 19, c, f"={col(c)}8*{AS}!{col(c)}31", fmt=num_fmt)

sl(ws10b, 20, 1, "Occupancy")
for c in range(3, 8):
    sc(ws10b, 20, c, f"={col(c)}8*{AS}!{col(c)}32", fmt=num_fmt)

sl(ws10b, 21, 1, "Gross profit", bold=True)
for c in range(3, 8):
    sc(ws10b, 21, c, f"={col(c)}14-{col(c)}17-{col(c)}18-{col(c)}19-{col(c)}20", fmt=num_fmt, bold=True)

sl(ws10b, 22, 1, "  Percentage margin")
for c in range(3, 8):
    sc(ws10b, 22, c, f"={col(c)}21/{col(c)}14", fmt=pct_fmt)

sl(ws10b, 23, 1, "SG&A")
for c in range(3, 8):
    sc(ws10b, 23, c, f"={col(c)}14*{AS}!{col(c)}33", fmt=num_fmt)

sl(ws10b, 24, 1, "EBITDA", bold=True)
sx(ws10b, 24, 2, f"={IS}!G28+{IS}!G25", fmt=dollar_fmt, bold=True)
for c in range(3, 8):
    sc(ws10b, 24, c, f"={col(c)}21-{col(c)}23", fmt=dollar_fmt, bold=True)

sl(ws10b, 25, 1, "  Percentage margin")
sc(ws10b, 25, 2, "=B24/B14", fmt=pct_fmt)
for c in range(3, 8):
    sc(ws10b, 25, c, f"={col(c)}24/{col(c)}14", fmt=pct_fmt)

sl(ws10b, 27, 1, "Depreciation and amortization")
for c in range(3, 8):
    sc(ws10b, 27, c, f"={AS}!{col(c)}35*{AS}!{col(c)}20", fmt=num_fmt)

sl(ws10b, 28, 1, "Operating income (EBIT)", bold=True)
sx(ws10b, 28, 2, f"={IS}!G28", fmt=dollar_fmt, bold=True)
for c in range(3, 8):
    sc(ws10b, 28, c, f"={col(c)}24-{col(c)}27", fmt=dollar_fmt, bold=True)

sl(ws10b, 29, 1, "  Percentage margin")
sc(ws10b, 29, 2, "=B28/B14", fmt=pct_fmt)
for c in range(3, 8):
    sc(ws10b, 29, c, f"={col(c)}28/{col(c)}14", fmt=pct_fmt)

# ============================================================
# SHEET 8: Exhibit 11 - Comps (ALL INPUT)
# ============================================================
ws11 = wb.create_sheet("Ex11 Comps")
ws11.column_dimensions['A'].width = 28
for c_idx, w in zip(range(2, 10), [16, 14, 14, 14, 14, 14, 14, 14]):
    ws11.column_dimensions[col(c_idx)].width = w

sl(ws11, 1, 1, "Exhibit 11 — Restaurant Comparable Valuation Analysis", section=True)
sl(ws11, 2, 1, "As of Oct. 13, 2017 | $ in millions | Source: Page 21")

comp_h = ["Company", "Mkt Cap", "% Fran.", "% Unit Gr.", "Sales Gr. CY16", "Sales Gr. CY17E", "EBITDA Mgn LTM", "P/EPS LTM", "TEV/EBITDA LTM"]

def write_comps(ws, start_row, header_row, data):
    set_headers(ws, header_row, comp_h)
    for i, rd in enumerate(data):
        r = start_row + i
        name = rd[0]
        is_bww = name == "Buffalo Wild Wings"
        f = Font(bold=True, color='0000FF') if is_bww else input_font
        ws.cell(row=r, column=1, value=name).font = Font(bold=True) if is_bww else Font()
        for j, (v, nf) in enumerate(zip(rd[1:], [dollar_int, '0%', '0%', '0%', '0%', pct_fmt, mult_fmt, mult_fmt])):
            cell = ws.cell(row=r, column=2+j, value=v)
            cell.font = f
            if not isinstance(v, str):
                cell.number_format = nf

# Casual Dining
sl(ws11, 4, 1, "Casual Dining", section=True)
casual = [
    ("Darden", 9822, 0, 0.12, 0.01, 0.09, 0.129, 20.4, 10.9),
    ("Cracker Barrel", 3645, 0, 0.01, 0.03, 0.01, 0.137, 18.1, 9.7),
    ("Texas Roadhouse", 3548, 0.16, 0.07, 0.10, 0.11, 0.133, 30.1, 12.5),
    ("Cheesecake Factory", 1983, 0, 0.04, 0.08, -0.01, 0.118, 14.8, 7.1),
    ("Bloomin' Brands", 1687, 0.20, -0.01, -0.03, -0.02, 0.097, 19.5, 6.9),
    ("Buffalo Wild Wings", 1587, 0.51, 0.05, 0.10, 0.04, 0.128, 25.9, 7.7),
    ("Brinker", 1520, 0.40, 0.01, 0.04, -0.03, 0.137, 10.7, 6.5),
    ("BJ's Restaurant", 645, 0, 0.05, 0.08, 0.04, 0.113, 17.7, 6.5),
    ("Red Robin", 828, 0.16, 0.01, 0.03, 0.06, 0.101, 99.3, 8.3),
    ("Chuy's", 350, 0, 0.16, 0.15, 0.12, 0.110, 20.9, 8.0),
]
write_comps(ws11, 6, 5, casual)

# Quick Service
sl(ws11, 17, 1, "Quick Service", section=True)
qsr = [
    ("McDonald's", 133949, 0.93, 0.01, -0.03, -0.08, 0.409, 27.1, 16.4),
    ("Yum Brands", 26272, 0.96, 0.03, -0.51, -0.07, 0.336, 28.8, 17.3),
    ("Restaurant Brands Int.", 15658, 1.00, 0.06, 0.02, 0.12, 0.446, 46.4, 15.2),
    ("The Wendy's Company", 3737, 0.95, 0.01, -0.23, -0.15, 0.312, 39.8, 15.9),
    ("Wingstop", 945, 0.98, 0.15, 0.17, 0.15, 0.284, 50.7, 30.8),
]
write_comps(ws11, 19, 18, qsr)

# Fast Casual
sl(ws11, 25, 1, "Fast Casual", section=True)
fast_casual = [
    ("Chipotle", 9308, 0, 0.08, -0.13, 0.15, 0.101, 70.2, 22.0),
    ("Shake Shack", 876, 0.42, 0.32, 0.41, 0.32, 0.158, 58.5, 17.2),
    ("Potbelly", 307, 0.12, 0.11, 0.09, 0.05, 0.088, 76.6, 7.0),
    ("Zoe's Kitchen", 229, 0.01, 0.19, 0.22, 0, 0.075, "NM", 11.8),
]
write_comps(ws11, 27, 26, fast_casual)

# Coffee
sl(ws11, 32, 1, "Coffee", section=True)
coffee = [
    ("Starbucks", 80454, 0.50, 0.09, 0.10, 0.05, 0.222, 28.1, 14.7),
    ("Dunkin' Brands", 5006, 1.00, 0.02, 0.06, 0.04, 0.371, 24.3, 15.4),
]
write_comps(ws11, 34, 33, coffee)

# ============================================================
# SHEET 9: Exhibit 12 - Debt Schedule
# ============================================================
ws12 = wb.create_sheet("Ex12 Debt")
ws12.column_dimensions['A'].width = 35
for c in range(2, 6):
    ws12.column_dimensions[col(c)].width = 16

sl(ws12, 1, 1, "Exhibit 12 — Debt Schedule", section=True)
sl(ws12, 2, 1, "At closing (expected Dec. 30, 2017) | $ in millions | Source: Page 22")
set_headers(ws12, 3, ["Instrument", "Amount", "LIBOR (Oct 2017)", "Spread", "Rate"])

sl(ws12, 4, 1, "Revolver (max $150M)")
si(ws12, 4, 2, 100, fmt=dollar_int)
si(ws12, 4, 3, 0.018, fmt=pct_fmt)
si(ws12, 4, 4, 0.035, fmt=pct_fmt)
sc(ws12, 4, 5, "=C4+D4", fmt=pct_fmt)

sl(ws12, 5, 1, "Senior Secured Term Loan B")
si(ws12, 5, 2, 1575, fmt=dollar_int)
si(ws12, 5, 3, 0.018, fmt=pct_fmt)
si(ws12, 5, 4, 0.035, fmt=pct_fmt)
sc(ws12, 5, 5, "=C5+D5", fmt=pct_fmt)

sl(ws12, 6, 1, "Senior unsecured notes")
si(ws12, 6, 2, 485, fmt=dollar_int)
ws12.cell(row=6, column=3, value="—")
ws12.cell(row=6, column=4, value="—")
si(ws12, 6, 5, 0.073, fmt=pct_fmt)

sl(ws12, 8, 1, "Total debt", bold=True)
sc(ws12, 8, 2, "=SUM(B4:B6)", fmt=dollar_int, bold=True)

# ============================================================
# SHEET 10: Exhibit 13 - Selected Transactions (ALL INPUT)
# ============================================================
ws13 = wb.create_sheet("Ex13 Transactions")
ws13.column_dimensions['A'].width = 8
ws13.column_dimensions['B'].width = 38
ws13.column_dimensions['C'].width = 18
ws13.column_dimensions['D'].width = 38
ws13.column_dimensions['E'].width = 16
ws13.column_dimensions['F'].width = 16

sl(ws13, 1, 1, "Exhibit 13 — Selected Transactions Analysis", section=True)
sl(ws13, 2, 1, "2011-2017 | $ in millions | Source: Page 23")
set_headers(ws13, 3, ["Date", "Target", "Type", "Acquirer", "Value ($M)", "TEV/EBITDA"])

txns = [
    (2017, "Ruby Tuesday", "Casual dining", "NRD Capital", 335, 7.9),
    (2017, "Panera Bread Company", "Casual dining", "JAB Holdings B.V.", 7500, 18.5),
    (2017, "Cheddar's Restaurant Holding", "Casual dining", "Darden Restaurants, Inc.", 780, 10.8),
    (2017, "Bob Evans", "Casual dining", "Golden Gate Capital", 565, 6.8),
    (2017, "Checkers Drive-In", "QSR/fast casual", "Oak Hill Capital Partners", None, 11.0),
    (2017, "Cheddar's Scratch Kitchen", "Casual dining", "Darden Restaurants", 780, 10.4),
    (2017, "Popeye's Louisiana Kitchen", "QSR/fast casual", "Restaurant Brands Int.", 1800, 20.6),
    (2016, "Krispy Kreme Donuts", "QSR/fast casual", "JAB Holdings B.V.", 1350, 19.0),
    (2015, "Del Taco Restaurants", "QSR/fast casual", "Levy Acquisition Corp.", 15500, 8.7),
    (2014, "Tim Horton's", "QSR/fast casual", "Burger King World Wide", 11160, 14.3),
    (2014, "Einstein's Noah Restaurant Group", "QSR/fast casual", "JAB Holdings B.V.", 374, 9.0),
    (2014, "Red Lobster", "Casual dining", "Golden Gate Capital", 2100, 9.2),
    (2014, "CEC Entertainment", "Casual dining", "Apollo Global Management", 1330, 7.9),
    (2014, "TGI Fridays", "Casual dining", "Sentinel/TriArtisan", 890, 7.5),
    (2013, "CKE", "QSR/fast casual", "Roark Capital", 1700, 8.3),
    (2014, "Caribou Coffee", "QSR/fast casual", "JAB Holdings B.V.", 340, 11.0),
    (2012, "Yard House USA", "Casual dining", "Darden Restaurants, Inc.", 585, 15.0),
    (2012, "P.F. Chang's China Bistro", "Casual dining", "Centerbridge Partners", 1100, 8.7),
    (2011, "California Pizza Kitchen", "Casual dining", "Golden Gate Capital", 470, 7.6),
    (2011, "Arby's", "QSR/fast casual", "Roark Capital", 340, 6.5),
]

for i, (date, target, typ, acquirer, value, mult) in enumerate(txns):
    r = 4 + i
    si(ws13, r, 1, date, fmt='0')
    si(ws13, r, 2, target)
    si(ws13, r, 3, typ)
    si(ws13, r, 4, acquirer)
    if value is not None:
        si(ws13, r, 5, value, fmt=dollar_int)
    else:
        si(ws13, r, 5, "—")
    si(ws13, r, 6, mult, fmt=mult_fmt)

# ============================================================
# SAVE
# ============================================================
output_path = "/Users/rj/Library/CloudStorage/OneDrive-MS365Workspace/5 DESK (OneDrive)/AWS GTHB/Client Repos/Alterrra-Capital-Client-Repo/New Course Material/SPREADSHEET RECREATION/Execution/260314 Test 1/03_Financial_Model.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
print(f"Sheets: {wb.sheetnames}")
